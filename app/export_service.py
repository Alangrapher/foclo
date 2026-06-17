"""Export timesheet to Excel / Markdown / JSON.

Excel: template format with SUM formulas, deduped descriptions.
Markdown: Obsidian-friendly with frontmatter, two tables (by Subject + by Date), wikilinks.
JSON: structured data for Dataview / external processing.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from collections import defaultdict
import json


DAY_ABBR = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]


def _week_range(ref_date: date | None = None) -> tuple[date, date]:
    d = ref_date or date.today()
    sunday = d - timedelta(days=(d.weekday() + 1) % 7)
    if sunday > d:
        sunday -= timedelta(days=7)
    return sunday, sunday + timedelta(days=6)


def _collect_data(start_date: date | None, end_date: date | None) -> dict:
    """Query DB and aggregate into a shared data structure.

    Returns dict with keys:
      s, e, subject_map, ordered_ids, all_day_hours, all_descriptions,
      num_days, day_labels, day_dates, total_hours, subject_names
    """
    from app.storage import get_conn
    from app.subject_service import get_subjects

    sweek, eweek = _week_range(start_date)
    s = start_date if start_date else sweek
    e = end_date if end_date else (start_date if start_date else eweek)

    # ── Week-range guard ──
    week_start = _week_range(s)[0]
    week_end = week_start + timedelta(days=6)
    if s < week_start or e > week_end:
        raise ValueError(
            f"Date range {s} → {e} spans more than one week.\n"
            f"This template supports a single Sun–Sat week only.\n"
            f"Narrow your selection to within {week_start} → {week_end}."
        )

    conn = get_conn()
    rows = conn.execute(
        """SELECT r.id, r.subject_id, r.description, r.start_time, r.duration_s,
                  s.name AS subject_name
           FROM records r
           LEFT JOIN subjects s ON r.subject_id = s.id
           WHERE date(r.start_time) BETWEEN ? AND ?
           ORDER BY r.subject_id, r.start_time""",
        (s.isoformat(), e.isoformat()),
    ).fetchall()
    conn.close()

    if not rows:
        raise ValueError(f"No records found for {s} – {e}")

    subject_groups: dict[int, list] = defaultdict(list)
    for r in rows:
        subject_groups[r["subject_id"] or 0].append(r)

    subjects = get_subjects()
    subject_map = {s["id"]: s for s in subjects}
    conn2 = get_conn()
    try:
        archived_rows = conn2.execute(
            "SELECT id, name FROM subjects WHERE archived = 1 AND id IN ("
            "SELECT DISTINCT subject_id FROM records WHERE subject_id IS NOT NULL)"
        ).fetchall()
    finally:
        conn2.close()
    for r in archived_rows:
        if r["id"] not in subject_map:
            subject_map[r["id"]] = dict(r)
    ordered_ids = list(subject_groups.keys())
    for subj in subjects:
        if subj["id"] not in ordered_ids:
            ordered_ids.append(subj["id"])

    all_day_hours: dict[int, dict[int, float]] = {}
    all_descriptions: dict[int, list[str]] = {}
    has_saturday = False

    for sid in ordered_ids:
        recs = subject_groups.get(sid, [])
        day_hours = {i: 0.0 for i in range(7)}
        desc_hours: dict[str, float] = {}

        for r in recs:
            dt_str = r["start_time"]
            dt = datetime.fromisoformat(dt_str) if isinstance(dt_str, str) else dt_str
            day_idx = (dt.weekday() + 1) % 7
            hours = (r["duration_s"] or 0) / 3600.0
            day_hours[day_idx] += hours
            if day_idx == 6:
                has_saturday = True
            desc = (r["description"] or "").strip()
            if desc:
                desc_hours[desc] = desc_hours.get(desc, 0) + hours

        descriptions = []
        for i, (desc, hrs) in enumerate(desc_hours.items(), 1):
            descriptions.append(f"{i}. {desc} ({hrs:.1f}h)")

        all_day_hours[sid] = day_hours
        all_descriptions[sid] = descriptions

    num_days = 7 if has_saturday else 6
    anchor_ref = start_date if start_date else date.today()
    anchor_sunday = anchor_ref - timedelta(days=(anchor_ref.weekday() + 1) % 7)
    if anchor_sunday > anchor_ref:
        anchor_sunday -= timedelta(days=7)

    day_dates = [anchor_sunday + timedelta(days=i) for i in range(num_days)]
    day_labels = [
        f"{DAY_ABBR[i]} {day_dates[i]:%Y%m%d}"
        for i in range(num_days)
    ]

    # Compute total hours
    total_hours = sum(
        sum(dh.values()) for dh in all_day_hours.values()
    )

    # Subject names (with wikilink-friendly keys)
    subject_names = {}
    for sid in ordered_ids:
        name = subject_map.get(sid, {}).get("name", f"Subject {sid}") if sid else "Uncategorised"
        subject_names[sid] = name

    return {
        "s": s, "e": e,
        "subject_map": subject_map,
        "ordered_ids": ordered_ids,
        "all_day_hours": all_day_hours,
        "all_descriptions": all_descriptions,
        "num_days": num_days,
        "day_labels": day_labels,
        "day_dates": day_dates,
        "total_hours": total_hours,
        "subject_names": subject_names,
    }


# ═══════════════════════════════════════════════════════
#  Excel
# ═══════════════════════════════════════════════════════

def export_xlsx(
    output_path: str,
    start_date: date | None = None,
    end_date: date | None = None,
) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return None  # caller handles: raise RuntimeError("openpyxl is not installed")

    data = _collect_data(start_date, end_date)
    s, e = data["s"], data["e"]
    ordered_ids = data["ordered_ids"]
    subject_map = data["subject_map"]
    subject_names = data["subject_names"]
    all_day_hours = data["all_day_hours"]
    all_descriptions = data["all_descriptions"]
    num_days = data["num_days"]
    day_labels = data["day_labels"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    header_font = Font(name="Calibri", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16

    headers = ["Project No.", "Project", "Descriptions (if any)",
               "Hour(s) - Weekly", "Days", "Hour(s) - Daily"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 2
    project_no = 0

    for sid in ordered_ids:
        project_no += 1
        name = subject_names[sid]
        day_hours = all_day_hours[sid]
        desc_text = "\n".join(all_descriptions.get(sid, []))

        day_row_start = row
        day_row_end = row + num_days - 1

        ws.cell(row=row, column=1, value=project_no).border = thin_border
        ws.cell(row=row, column=2, value=name).border = thin_border
        desc_cell = ws.cell(row=row, column=3, value=desc_text)
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

        formula_cell = ws.cell(row=row, column=4,
                               value=f"=SUM(F{day_row_start}:F{day_row_end})")
        formula_cell.border = thin_border
        formula_cell.alignment = Alignment(horizontal="center", vertical="center")
        formula_cell.number_format = "0.0"

        ws.cell(row=row, column=5, value=day_labels[0]).border = thin_border
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
        h0 = ws.cell(row=row, column=6, value=round(day_hours[0], 2))
        h0.border = thin_border
        h0.alignment = Alignment(horizontal="center", vertical="center")
        h0.number_format = "0.0"

        for day_idx in range(1, num_days):
            r_num = row + day_idx
            for col in (1, 2, 3, 4):
                ws.cell(row=r_num, column=col).border = thin_border
            ws.cell(row=r_num, column=5, value=day_labels[day_idx]).border = thin_border
            ws.cell(row=r_num, column=5).alignment = Alignment(horizontal="center", vertical="center")
            h_cell = ws.cell(row=r_num, column=6, value=round(day_hours[day_idx], 2))
            h_cell.border = thin_border
            h_cell.alignment = Alignment(horizontal="center", vertical="center")
            h_cell.number_format = "0.0"

        row = row + num_days

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════
#  Markdown
# ═══════════════════════════════════════════════════════

def export_markdown(
    output_path: str,
    start_date: date | None = None,
    end_date: date | None = None,
) -> str:
    data = _collect_data(start_date, end_date)
    s, e = data["s"], data["e"]
    ordered_ids = data["ordered_ids"]
    subject_names = data["subject_names"]
    all_day_hours = data["all_day_hours"]
    all_descriptions = data["all_descriptions"]
    num_days = data["num_days"]
    day_labels = data["day_labels"]
    day_dates = data["day_dates"]
    total_hours = data["total_hours"]

    lines = []

    # Frontmatter
    subjects_list = [subject_names[sid] for sid in ordered_ids]
    lines.append("---")
    lines.append(f"date: {s.isoformat()}")
    lines.append(f"end_date: {e.isoformat()}")
    lines.append("type: timesheet")
    lines.append(f"subjects: [{', '.join(subjects_list)}]")
    lines.append(f"total_hours: {total_hours:.1f}")
    lines.append("---")
    lines.append("")

    # Title
    lines.append(f"# Timesheet {s.isoformat()} → {e.isoformat()}")
    lines.append("")

    # ── By Subject ──
    lines.append("## By Subject")
    lines.append("")

    for sid in ordered_ids:
        name = subject_names[sid]
        day_hours = all_day_hours[sid]
        descs = all_descriptions.get(sid, [])
        subject_total = sum(day_hours.values())

        # Wikilink heading
        lines.append(f"### [[{name}]]  — {subject_total:.1f}h")
        lines.append("")
        lines.append("| Day | Hours | Description |")
        lines.append("|-----|-------|-------------|")

        for day_idx in range(num_days):
            h = day_hours.get(day_idx, 0)
            label_short = f"{DAY_ABBR[day_idx]} {day_dates[day_idx]:%m-%d}"
            h_str = f"{h:.1f}" if h > 0 else "—"
            lines.append(f"| {label_short} | {h_str} | |")

        # Description summary line
        if descs:
            desc_joined = " / ".join(descs)
            desc_joined = desc_joined.replace("|", "\\|")
            lines.append(f"| | | {desc_joined} |")
        lines.append("")

    # ── By Date ──
    lines.append("## By Date")
    lines.append("")

    for day_idx in range(num_days):
        label = day_labels[day_idx]
        day_total = 0.0
        day_rows = []

        for sid in ordered_ids:
            h = all_day_hours[sid].get(day_idx, 0)
            if h > 0:
                name = subject_names[sid]
                day_total += h
                day_rows.append((name, h))

        if not day_rows:
            continue

        lines.append(f"### {day_dates[day_idx]:%a %Y-%m-%d}  — {day_total:.1f}h")
        lines.append("")
        lines.append("| Subject | Hours | Description |")
        lines.append("|---------|-------|-------------|")

        for name, h in day_rows:
            lines.append(f"| [[{name}]] | {h:.1f} | |")

        lines.append("")

    md = "\n".join(lines) + "\n"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md)
    return output_path


# ═══════════════════════════════════════════════════════
#  JSON
# ═══════════════════════════════════════════════════════

def export_json(
    output_path: str,
    start_date: date | None = None,
    end_date: date | None = None,
) -> str:
    data = _collect_data(start_date, end_date)
    s, e = data["s"], data["e"]
    ordered_ids = data["ordered_ids"]
    subject_names = data["subject_names"]
    all_day_hours = data["all_day_hours"]
    all_descriptions = data["all_descriptions"]
    num_days = data["num_days"]
    day_labels = data["day_labels"]
    day_dates = data["day_dates"]
    total_hours = data["total_hours"]

    # By subject
    by_subject = []
    for sid in ordered_ids:
        day_hours = all_day_hours[sid]
        s_total = sum(day_hours.values())
        days = []
        for day_idx in range(num_days):
            days.append({
                "day": day_labels[day_idx],
                "date": day_dates[day_idx].isoformat(),
                "hours": round(day_hours[day_idx], 2),
            })
        by_subject.append({
            "subject": subject_names[sid],
            "total_hours": round(s_total, 2),
            "descriptions": all_descriptions.get(sid, []),
            "days": days,
        })

    # By date
    by_date = []
    for day_idx in range(num_days):
        entries = []
        day_total = 0.0
        for sid in ordered_ids:
            h = all_day_hours[sid].get(day_idx, 0)
            if h > 0:
                entries.append({"subject": subject_names[sid], "hours": round(h, 2)})
                day_total += h
        by_date.append({
            "date": day_dates[day_idx].isoformat(),
            "day": day_labels[day_idx],
            "total_hours": round(day_total, 2),
            "entries": entries,
        })

    out = {
        "start_date": s.isoformat(),
        "end_date": e.isoformat(),
        "total_hours": round(total_hours, 2),
        "by_subject": by_subject,
        "by_date": by_date,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    return output_path


# ═══════════════════════════════════════════════════════
#  Unified entry point (keeps backward compat)
# ═══════════════════════════════════════════════════════

def export_timesheet(
    output_path: str,
    start_date: date | None = None,
    end_date: date | None = None,
) -> str:
    """Backward-compatible shortcut — always exports Excel."""
    return export_xlsx(output_path, start_date, end_date)
